"""
Script 2: Search Patterns Analysis
Analyzes structured fields (shoot_types, usages, locations, etc.) and their combinations
"""

import sys
import subprocess
import os

# Check and install required packages
required_packages = {
    'pandas': 'pandas',
    'numpy': 'numpy',
    'matplotlib': 'matplotlib',
    'plotly': 'plotly',
    'seaborn': 'seaborn',
    'itertools': 'itertools'
}

print("Checking required packages...")
for package, pip_name in required_packages.items():
    if package == 'itertools':  # itertools is built-in
        continue
    try:
        __import__(package)
        print(f"✓ {package} already installed")
    except ImportError:
        print(f"✗ {package} not found. Installing {pip_name}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name, "--break-system-packages"])
        print(f"✓ {pip_name} installed successfully")

# Now import everything
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import seaborn as sns
from collections import Counter
from itertools import combinations
import json
import re

print("\n" + "="*80)
print("SCRIPT 2: SEARCH PATTERNS ANALYSIS")
print("="*80 + "\n")

# Configuration
DATA_DIR = "data"
OUTPUT_DIR = "outputs/patterns"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Read data
print("Loading data...")
df = pd.read_csv(f"{DATA_DIR}/2025_Bookings.csv")
print(f"✓ Loaded {len(df)} bookings from {df['job_id'].nunique()} unique jobs\n")

# Deduplicate at job level for pattern analysis
print("Deduplicating at job level...")
job_df = df.drop_duplicates(subset='job_id', keep='first')
print(f"✓ {len(job_df)} unique jobs for pattern analysis\n")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def parse_array_field(field_value):
    """Parse fields that are in {value1,value2} format"""
    if pd.isna(field_value):
        return []
    
    field_str = str(field_value).strip()
    
    # Handle empty arrays
    if field_str == '{}' or field_str == '':
        return []
    
    # Remove curly braces
    field_str = field_str.strip('{}')
    
    # Split by comma
    values = [v.strip().strip('"') for v in field_str.split(',')]
    
    return [v for v in values if v]

def get_combination_name(items):
    """Create readable name for combination"""
    return " + ".join(sorted(items))

# ============================================================================
# PARSE STRUCTURED FIELDS
# ============================================================================

print("Parsing structured fields...")

# Parse array fields
job_df['shoot_types_list'] = job_df['shoot_types'].apply(parse_array_field)
job_df['shoot_locations_list'] = job_df['shoot_locations'].apply(parse_array_field)
job_df['shoot_details_list'] = job_df['shoot_details'].apply(parse_array_field)
job_df['usages_list'] = job_df['usages'].apply(parse_array_field)
job_df['extra_needs_list'] = job_df['extra_needs'].apply(parse_array_field)

print("✓ Parsed all structured fields\n")

# ============================================================================
# INDIVIDUAL FIELD ANALYSIS
# ============================================================================

print("="*80)
print("ANALYZING INDIVIDUAL FIELDS")
print("="*80 + "\n")

# 1. Shoot Types
print("1. SHOOT TYPES")
print("-" * 40)
all_shoot_types = []
for types in job_df['shoot_types_list']:
    all_shoot_types.extend(types)

shoot_type_freq = Counter(all_shoot_types)
print(f"Total mentions: {len(all_shoot_types)}")
print(f"Unique types: {len(shoot_type_freq)}")
print("\nFrequency:")
for type_name, count in shoot_type_freq.most_common():
    pct = (count / len(job_df)) * 100
    print(f"  {type_name}: {count} ({pct:.1f}%)")
print()

# 2. Shoot Locations
print("2. SHOOT LOCATIONS")
print("-" * 40)
all_locations = []
for locs in job_df['shoot_locations_list']:
    all_locations.extend(locs)

location_freq = Counter(all_locations)
print(f"Total mentions: {len(all_locations)}")
print(f"Unique locations: {len(location_freq)}")
print("\nFrequency:")
for loc, count in location_freq.most_common():
    pct = (count / len(job_df)) * 100
    print(f"  {loc}: {count} ({pct:.1f}%)")
print()

# 3. Shoot Details
print("3. SHOOT DETAILS")
print("-" * 40)
all_details = []
for details in job_df['shoot_details_list']:
    all_details.extend(details)

detail_freq = Counter(all_details)
print(f"Total mentions: {len(all_details)}")
print(f"Unique details: {len(detail_freq)}")
print("\nTop 10 details:")
for detail, count in detail_freq.most_common(10):
    pct = (count / len(job_df)) * 100
    print(f"  {detail}: {count} ({pct:.1f}%)")
print()

# 4. Usages
print("4. USAGES")
print("-" * 40)
all_usages = []
for usages in job_df['usages_list']:
    all_usages.extend(usages)

usage_freq = Counter(all_usages)
print(f"Total mentions: {len(all_usages)}")
print(f"Unique usages: {len(usage_freq)}")
print("\nFrequency:")
for usage, count in usage_freq.most_common():
    pct = (count / len(job_df)) * 100
    print(f"  {usage}: {count} ({pct:.1f}%)")
print()

# 5. Copyright Duration
print("5. COPYRIGHT DURATION (months)")
print("-" * 40)
copyright_dist = job_df['copyright'].value_counts().sort_index()
print("Distribution:")
for months, count in copyright_dist.items():
    pct = (count / len(job_df)) * 100
    print(f"  {months} months: {count} ({pct:.1f}%)")
print(f"\nMean: {job_df['copyright'].mean():.1f} months")
print(f"Median: {job_df['copyright'].median():.0f} months")
print()

# 6. Shoot Hours
print("6. SHOOT HOURS")
print("-" * 40)
hours_dist = job_df['shoot_hours'].value_counts().sort_index()
print("Distribution:")
for hours, count in hours_dist.head(10).items():
    pct = (count / len(job_df)) * 100
    print(f"  {hours} hours: {count} ({pct:.1f}%)")
print(f"\nMean: {job_df['shoot_hours'].mean():.1f} hours")
print(f"Median: {job_df['shoot_hours'].median():.0f} hours")
print()

# 7. Extra Needs
print("7. EXTRA NEEDS")
print("-" * 40)
all_extra = []
for extra in job_df['extra_needs_list']:
    all_extra.extend(extra)

extra_freq = Counter(all_extra)
print(f"Total mentions: {len(all_extra)}")
print(f"Jobs with extra needs: {len([e for e in job_df['extra_needs_list'] if e])}")
if extra_freq:
    print("\nTop 10 extra needs:")
    for need, count in extra_freq.most_common(10):
        pct = (count / len(job_df)) * 100
        print(f"  {need}: {count} ({pct:.1f}%)")
print()

# ============================================================================
# COMBINATION ANALYSIS
# ============================================================================

print("="*80)
print("ANALYZING COMMON COMBINATIONS")
print("="*80 + "\n")

# 1. Shoot Type + Location combinations
print("1. SHOOT TYPE + LOCATION COMBINATIONS")
print("-" * 40)
type_location_combos = []
for idx, row in job_df.iterrows():
    types = row['shoot_types_list']
    locations = row['shoot_locations_list']
    for t in types:
        for l in locations:
            type_location_combos.append(f"{t} + {l}")

type_loc_freq = Counter(type_location_combos)
print("Top 15 combinations:")
for combo, count in type_loc_freq.most_common(15):
    pct = (count / len(job_df)) * 100
    print(f"  {combo}: {count} ({pct:.1f}%)")
print()

# 2. Shoot Type + Usage combinations
print("2. SHOOT TYPE + USAGE COMBINATIONS")
print("-" * 40)
type_usage_combos = []
for idx, row in job_df.iterrows():
    types = row['shoot_types_list']
    usages = row['usages_list']
    for t in types:
        for u in usages:
            type_usage_combos.append(f"{t} + {u}")

type_usage_freq = Counter(type_usage_combos)
print("Top 15 combinations:")
for combo, count in type_usage_freq.most_common(15):
    pct = (count / len(job_df)) * 100
    print(f"  {combo}: {count} ({pct:.1f}%)")
print()

# 3. Location + Usage combinations
print("3. LOCATION + USAGE COMBINATIONS")
print("-" * 40)
loc_usage_combos = []
for idx, row in job_df.iterrows():
    locations = row['shoot_locations_list']
    usages = row['usages_list']
    for l in locations:
        for u in usages:
            loc_usage_combos.append(f"{l} + {u}")

loc_usage_freq = Counter(loc_usage_combos)
print("Top 15 combinations:")
for combo, count in loc_usage_freq.most_common(15):
    pct = (count / len(job_df)) * 100
    print(f"  {combo}: {count} ({pct:.1f}%)")
print()

# 4. Full pattern: Type + Location + Usage
print("4. FULL PATTERN: TYPE + LOCATION + USAGE")
print("-" * 40)
full_patterns = []
for idx, row in job_df.iterrows():
    types = row['shoot_types_list']
    locations = row['shoot_locations_list']
    usages = row['usages_list']
    
    if types and locations and usages:
        # Take the first of each for simplicity
        pattern = f"{types[0]} + {locations[0]} + {usages[0]}"
        full_patterns.append(pattern)

full_pattern_freq = Counter(full_patterns)
print("Top 20 full patterns:")
for pattern, count in full_pattern_freq.most_common(20):
    pct = (count / len(job_df)) * 100
    print(f"  {pattern}: {count} ({pct:.1f}%)")
print()

# ============================================================================
# VISUALIZATIONS
# ============================================================================

print("="*80)
print("GENERATING VISUALIZATIONS")
print("="*80 + "\n")

# Set color scheme
colors = px.colors.qualitative.Set3

# 1. Shoot Types Distribution
print("1. Creating shoot types chart...")
fig = go.Figure(data=[
    go.Bar(
        x=list(shoot_type_freq.values()),
        y=list(shoot_type_freq.keys()),
        orientation='h',
        marker_color=colors[0]
    )
])
fig.update_layout(
    title='Shoot Types Distribution',
    xaxis_title='Frequency',
    yaxis_title='Shoot Type',
    height=400
)
fig.write_html(f'{OUTPUT_DIR}/shoot_types.html')
print(f"✓ Saved: {OUTPUT_DIR}/shoot_types.html")

# 2. Shoot Locations Distribution
print("2. Creating shoot locations chart...")
fig = go.Figure(data=[
    go.Bar(
        x=list(location_freq.values()),
        y=list(location_freq.keys()),
        orientation='h',
        marker_color=colors[1]
    )
])
fig.update_layout(
    title='Shoot Locations Distribution',
    xaxis_title='Frequency',
    yaxis_title='Location',
    height=400
)
fig.write_html(f'{OUTPUT_DIR}/shoot_locations.html')
print(f"✓ Saved: {OUTPUT_DIR}/shoot_locations.html")

# 3. Usages Distribution
print("3. Creating usages chart...")
fig = go.Figure(data=[
    go.Bar(
        x=list(usage_freq.values()),
        y=list(usage_freq.keys()),
        orientation='h',
        marker_color=colors[2]
    )
])
fig.update_layout(
    title='Usage Types Distribution',
    xaxis_title='Frequency',
    yaxis_title='Usage Type',
    height=500
)
fig.write_html(f'{OUTPUT_DIR}/usages.html')
print(f"✓ Saved: {OUTPUT_DIR}/usages.html")

# 4. Copyright Duration Distribution
print("4. Creating copyright duration chart...")
fig = go.Figure(data=[
    go.Bar(
        x=copyright_dist.index,
        y=copyright_dist.values,
        marker_color=colors[3]
    )
])
fig.update_layout(
    title='Copyright Duration Distribution',
    xaxis_title='Months',
    yaxis_title='Frequency',
    height=500
)
fig.write_html(f'{OUTPUT_DIR}/copyright_duration.html')
print(f"✓ Saved: {OUTPUT_DIR}/copyright_duration.html")

# 5. Shoot Hours Distribution
print("5. Creating shoot hours chart...")
fig = go.Figure(data=[
    go.Histogram(
        x=job_df['shoot_hours'],
        nbinsx=20,
        marker_color=colors[4]
    )
])
fig.update_layout(
    title='Shoot Hours Distribution',
    xaxis_title='Hours',
    yaxis_title='Frequency',
    height=500
)
fig.write_html(f'{OUTPUT_DIR}/shoot_hours.html')
print(f"✓ Saved: {OUTPUT_DIR}/shoot_hours.html")

# 6. Top Combinations - Type + Location
print("6. Creating type+location combinations chart...")
top_type_loc = dict(type_loc_freq.most_common(15))
fig = go.Figure(data=[
    go.Bar(
        x=list(top_type_loc.values()),
        y=list(top_type_loc.keys()),
        orientation='h',
        marker_color=colors[5]
    )
])
fig.update_layout(
    title='Top 15: Shoot Type + Location Combinations',
    xaxis_title='Frequency',
    yaxis_title='Combination',
    height=600,
    yaxis={'categoryorder': 'total ascending'}
)
fig.write_html(f'{OUTPUT_DIR}/combo_type_location.html')
print(f"✓ Saved: {OUTPUT_DIR}/combo_type_location.html")

# 7. Top Combinations - Type + Usage
print("7. Creating type+usage combinations chart...")
top_type_usage = dict(type_usage_freq.most_common(15))
fig = go.Figure(data=[
    go.Bar(
        x=list(top_type_usage.values()),
        y=list(top_type_usage.keys()),
        orientation='h',
        marker_color=colors[6]
    )
])
fig.update_layout(
    title='Top 15: Shoot Type + Usage Combinations',
    xaxis_title='Frequency',
    yaxis_title='Combination',
    height=600,
    yaxis={'categoryorder': 'total ascending'}
)
fig.write_html(f'{OUTPUT_DIR}/combo_type_usage.html')
print(f"✓ Saved: {OUTPUT_DIR}/combo_type_usage.html")

# 8. Heatmap - Type vs Location
print("8. Creating type vs location heatmap...")
# Create matrix
type_list = list(shoot_type_freq.keys())
loc_list = list(location_freq.keys())
matrix = np.zeros((len(type_list), len(loc_list)))

for idx, row in job_df.iterrows():
    for t in row['shoot_types_list']:
        for l in row['shoot_locations_list']:
            if t in type_list and l in loc_list:
                t_idx = type_list.index(t)
                l_idx = loc_list.index(l)
                matrix[t_idx][l_idx] += 1

fig = go.Figure(data=go.Heatmap(
    z=matrix,
    x=loc_list,
    y=type_list,
    colorscale='Blues'
))
fig.update_layout(
    title='Shoot Type vs Location Heatmap',
    xaxis_title='Location',
    yaxis_title='Shoot Type',
    height=500
)
fig.write_html(f'{OUTPUT_DIR}/heatmap_type_location.html')
print(f"✓ Saved: {OUTPUT_DIR}/heatmap_type_location.html")

# 9. Sunburst - Full Pattern Hierarchy
print("9. Creating pattern hierarchy sunburst chart...")
# Prepare data for sunburst
sunburst_data = []
for idx, row in job_df.iterrows():
    types = row['shoot_types_list']
    locations = row['shoot_locations_list']
    usages = row['usages_list']
    
    if types and locations and usages:
        sunburst_data.append({
            'type': types[0],
            'location': locations[0],
            'usage': usages[0]
        })

sunburst_df = pd.DataFrame(sunburst_data)

fig = go.Figure(go.Sunburst(
    labels=['All'] + 
           [f"Type: {t}" for t in shoot_type_freq.keys()] +
           [f"Loc: {l}" for l in location_freq.keys()] +
           [f"Usage: {u}" for u in usage_freq.keys()],
    parents=[''] + 
            ['All'] * len(shoot_type_freq) +
            ['All'] * len(location_freq) +
            ['All'] * len(usage_freq),
    values=[len(job_df)] +
           list(shoot_type_freq.values()) +
           list(location_freq.values()) +
           list(usage_freq.values())
))
fig.update_layout(
    title='Search Pattern Hierarchy',
    height=700
)
fig.write_html(f'{OUTPUT_DIR}/pattern_hierarchy.html')
print(f"✓ Saved: {OUTPUT_DIR}/pattern_hierarchy.html")

# ============================================================================
# EXPORT DATA
# ============================================================================

print("\n" + "="*80)
print("EXPORTING DATA")
print("="*80 + "\n")

# 1. Individual field frequencies
field_stats = {
    'shoot_types': shoot_type_freq,
    'shoot_locations': location_freq,
    'shoot_details': detail_freq,
    'usages': usage_freq,
    'extra_needs': extra_freq
}

for field_name, freq_dict in field_stats.items():
    df_export = pd.DataFrame([
        {'Value': k, 'Frequency': v, 'Percentage': (v/len(job_df))*100}
        for k, v in freq_dict.most_common()
    ])
    df_export.to_csv(f'{OUTPUT_DIR}/{field_name}_frequency.csv', index=False, encoding='utf-8-sig')
    print(f"✓ Saved: {OUTPUT_DIR}/{field_name}_frequency.csv")

# 2. Combination frequencies
combo_data = {
    'type_location': type_loc_freq,
    'type_usage': type_usage_freq,
    'location_usage': loc_usage_freq,
    'full_pattern': full_pattern_freq
}

for combo_name, freq_dict in combo_data.items():
    df_export = pd.DataFrame([
        {'Combination': k, 'Frequency': v, 'Percentage': (v/len(job_df))*100}
        for k, v in freq_dict.most_common(50)
    ])
    df_export.to_csv(f'{OUTPUT_DIR}/{combo_name}_combinations.csv', index=False, encoding='utf-8-sig')
    print(f"✓ Saved: {OUTPUT_DIR}/{combo_name}_combinations.csv")

# 3. Copyright and hours statistics
stats_df = pd.DataFrame({
    'Metric': ['Copyright (months)', 'Shoot Hours'],
    'Mean': [job_df['copyright'].mean(), job_df['shoot_hours'].mean()],
    'Median': [job_df['copyright'].median(), job_df['shoot_hours'].median()],
    'Min': [job_df['copyright'].min(), job_df['shoot_hours'].min()],
    'Max': [job_df['copyright'].max(), job_df['shoot_hours'].max()]
})
stats_df.to_csv(f'{OUTPUT_DIR}/numeric_statistics.csv', index=False, encoding='utf-8-sig')
print(f"✓ Saved: {OUTPUT_DIR}/numeric_statistics.csv")

# 4. Full parsed data
export_cols = ['job_id', 'brand_name', 'job_name', 'region', 
               'shoot_types_list', 'shoot_locations_list', 'shoot_details_list',
               'usages_list', 'extra_needs_list', 'shoot_hours', 'copyright']
job_df[export_cols].to_csv(f'{OUTPUT_DIR}/parsed_patterns_data.csv', index=False, encoding='utf-8-sig')
print(f"✓ Saved: {OUTPUT_DIR}/parsed_patterns_data.csv")

# ============================================================================
# GENERATE HTML REPORT
# ============================================================================

print("\n" + "="*80)
print("GENERATING HTML REPORT")
print("="*80 + "\n")

html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Search Patterns Analysis Report</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
            max-width: 1400px;
            margin: 0 auto;
            padding: 40px;
            background-color: #f5f5f5;
        }}
        .header {{
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
            color: white;
            padding: 40px;
            border-radius: 10px;
            margin-bottom: 30px;
        }}
        h1 {{
            margin: 0;
            font-size: 32px;
        }}
        .subtitle {{
            margin-top: 10px;
            opacity: 0.9;
        }}
        .section {{
            background: white;
            padding: 30px;
            margin-bottom: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h2 {{
            color: #11998e;
            border-bottom: 2px solid #11998e;
            padding-bottom: 10px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-box {{
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-number {{
            font-size: 36px;
            font-weight: bold;
        }}
        .stat-label {{
            font-size: 14px;
            opacity: 0.9;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #11998e;
            color: white;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .chart-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .chart-link {{
            padding: 20px;
            background: #f8f9fa;
            border-radius: 8px;
            text-align: center;
            text-decoration: none;
            color: #11998e;
            border: 2px solid #11998e;
            transition: all 0.3s;
        }}
        .chart-link:hover {{
            background: #11998e;
            color: white;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🔍 Search Patterns Analysis Report</h1>
        <div class="subtitle">Analysis of shoot types, locations, usages, and their combinations</div>
    </div>

    <div class="section">
        <h2>📈 Overview Statistics</h2>
        <div class="stats">
            <div class="stat-box">
                <div class="stat-number">{len(job_df)}</div>
                <div class="stat-label">Unique Jobs Analyzed</div>
            </div>
            <div class="stat-box">
                <div class="stat-number">{len(shoot_type_freq)}</div>
                <div class="stat-label">Shoot Types</div>
            </div>
            <div class="stat-box">
                <div class="stat-number">{len(location_freq)}</div>
                <div class="stat-label">Location Types</div>
            </div>
            <div class="stat-box">
                <div class="stat-number">{len(usage_freq)}</div>
                <div class="stat-label">Usage Types</div>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>📸 Shoot Types Distribution</h2>
        <table>
            <thead>
                <tr>
                    <th>Shoot Type</th>
                    <th>Frequency</th>
                    <th>Percentage</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"<tr><td>{t}</td><td>{c}</td><td>{(c/len(job_df))*100:.1f}%</td></tr>" 
                          for t, c in shoot_type_freq.most_common()])}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>📍 Shoot Locations Distribution</h2>
        <table>
            <thead>
                <tr>
                    <th>Location</th>
                    <th>Frequency</th>
                    <th>Percentage</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"<tr><td>{l}</td><td>{c}</td><td>{(c/len(job_df))*100:.1f}%</td></tr>" 
                          for l, c in location_freq.most_common()])}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>💼 Usage Types Distribution</h2>
        <table>
            <thead>
                <tr>
                    <th>Usage Type</th>
                    <th>Frequency</th>
                    <th>Percentage</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"<tr><td>{u}</td><td>{c}</td><td>{(c/len(job_df))*100:.1f}%</td></tr>" 
                          for u, c in usage_freq.most_common()])}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>🔗 Top Combinations</h2>
        <h3>Type + Location</h3>
        <table>
            <thead>
                <tr>
                    <th>Rank</th>
                    <th>Combination</th>
                    <th>Frequency</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"<tr><td>{i+1}</td><td>{combo}</td><td>{count}</td></tr>" 
                          for i, (combo, count) in enumerate(type_loc_freq.most_common(15))])}
            </tbody>
        </table>

        <h3>Type + Usage</h3>
        <table>
            <thead>
                <tr>
                    <th>Rank</th>
                    <th>Combination</th>
                    <th>Frequency</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"<tr><td>{i+1}</td><td>{combo}</td><td>{count}</td></tr>" 
                          for i, (combo, count) in enumerate(type_usage_freq.most_common(15))])}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>📊 Interactive Charts</h2>
        <div class="chart-grid">
            <a href="shoot_types.html" class="chart-link" target="_blank">
                📸 Shoot Types Chart
            </a>
            <a href="shoot_locations.html" class="chart-link" target="_blank">
                📍 Locations Chart
            </a>
            <a href="usages.html" class="chart-link" target="_blank">
                💼 Usages Chart
            </a>
            <a href="copyright_duration.html" class="chart-link" target="_blank">
                ⏱️ Copyright Duration
            </a>
            <a href="shoot_hours.html" class="chart-link" target="_blank">
                ⏰ Shoot Hours
            </a>
            <a href="combo_type_location.html" class="chart-link" target="_blank">
                🔗 Type + Location
            </a>
            <a href="combo_type_usage.html" class="chart-link" target="_blank">
                🔗 Type + Usage
            </a>
            <a href="heatmap_type_location.html" class="chart-link" target="_blank">
                🔥 Type vs Location Heatmap
            </a>
            <a href="pattern_hierarchy.html" class="chart-link" target="_blank">
                🌐 Pattern Hierarchy
            </a>
        </div>
    </div>

    <div class="section">
        <h2>📁 Exported Files</h2>
        <ul>
            <li><strong>shoot_types_frequency.csv</strong> - Shoot type frequencies</li>
            <li><strong>shoot_locations_frequency.csv</strong> - Location frequencies</li>
            <li><strong>usages_frequency.csv</strong> - Usage type frequencies</li>
            <li><strong>type_location_combinations.csv</strong> - Type + Location combos</li>
            <li><strong>type_usage_combinations.csv</strong> - Type + Usage combos</li>
            <li><strong>full_pattern_combinations.csv</strong> - Complete patterns</li>
            <li><strong>numeric_statistics.csv</strong> - Copyright and hours stats</li>
            <li><strong>parsed_patterns_data.csv</strong> - Full parsed dataset</li>
        </ul>
    </div>

    <div class="section">
        <h2>💡 Key Insights</h2>
        <ul>
            <li><strong>Most common shoot type:</strong> {shoot_type_freq.most_common(1)[0][0]} ({shoot_type_freq.most_common(1)[0][1]} jobs)</li>
            <li><strong>Most common location:</strong> {location_freq.most_common(1)[0][0]} ({location_freq.most_common(1)[0][1]} jobs)</li>
            <li><strong>Most common usage:</strong> {usage_freq.most_common(1)[0][0]} ({usage_freq.most_common(1)[0][1]} jobs)</li>
            <li><strong>Average copyright duration:</strong> {job_df['copyright'].mean():.1f} months</li>
            <li><strong>Average shoot hours:</strong> {job_df['shoot_hours'].mean():.1f} hours</li>
            <li><strong>Top combination:</strong> {type_loc_freq.most_common(1)[0][0]} ({type_loc_freq.most_common(1)[0][1]} jobs)</li>
        </ul>
    </div>
</body>
</html>
"""

with open(f'{OUTPUT_DIR}/patterns_report.html', 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"✓ Saved: {OUTPUT_DIR}/patterns_report.html")

# ============================================================================
# GENERATE EXCEL WORKBOOK
# ============================================================================

print("\nGenerating Excel workbook with all data...")

try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Add sheets for each dataset
    datasets = {
        'Shoot Types': pd.DataFrame([{'Value': k, 'Frequency': v, 'Percentage': f"{(v/len(job_df))*100:.1f}%"} 
                                      for k, v in shoot_type_freq.most_common()]),
        'Locations': pd.DataFrame([{'Value': k, 'Frequency': v, 'Percentage': f"{(v/len(job_df))*100:.1f}%"} 
                                    for k, v in location_freq.most_common()]),
        'Usages': pd.DataFrame([{'Value': k, 'Frequency': v, 'Percentage': f"{(v/len(job_df))*100:.1f}%"} 
                                 for k, v in usage_freq.most_common()]),
        'Type+Location': pd.DataFrame([{'Combination': k, 'Frequency': v} 
                                        for k, v in type_loc_freq.most_common(30)]),
        'Type+Usage': pd.DataFrame([{'Combination': k, 'Frequency': v} 
                                     for k, v in type_usage_freq.most_common(30)]),
        'Full Patterns': pd.DataFrame([{'Pattern': k, 'Frequency': v} 
                                        for k, v in full_pattern_freq.most_common(50)])
    }
    
    for sheet_name, df_data in datasets.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="11998e", end_color="11998e", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(f'{OUTPUT_DIR}/patterns_analysis.xlsx')
    print(f"✓ Saved: {OUTPUT_DIR}/patterns_analysis.xlsx")
    
except ImportError:
    print("⚠ openpyxl not installed. Skipping Excel export.")
    print("  Install with: pip install openpyxl")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "="*80)
print("✅ SCRIPT 2 COMPLETE!")
print("="*80 + "\n")

print("📊 Summary:")
print(f"  • Analyzed {len(job_df)} unique jobs")
print(f"  • Found {len(shoot_type_freq)} shoot types")
print(f"  • Found {len(location_freq)} location types")
print(f"  • Found {len(usage_freq)} usage types")
print(f"  • Identified {len(type_loc_freq)} unique combinations")
print(f"  • Generated {len([f for f in os.listdir(OUTPUT_DIR) if f.endswith(('.png', '.html', '.csv'))])} output files")
print(f"\n📁 All outputs saved to: {OUTPUT_DIR}/")
print(f"\n🌐 Open 'patterns_report.html' in your browser to view the interactive report!")
print("\n" + "="*80 + "\n")
